import sys
import os
import unicodedata
import csv
import numpy as np
import subprocess
from datetime import datetime, date, timedelta
import xlrd
import time
from openpyxl import load_workbook
from shutil import copy, copy2


def Leitura_Configuracao(endereco_conf):
    try:
        book = xlrd.open_workbook(endereco_conf)  # Procura o arquivo de configurações
    except:
        print("Favor, coloque o arquivo de configuracao SMAP em: " + endereco_conf)
        sys.exit()
    sheet_names = book.sheet_names()

    dictPosto_Distr = {}
    dictPosto_Montante_Viagem = {}
    for bacia in sheet_names:
        sheet = book.sheet_by_name(bacia)
        sessao = -1  # Cada sub-bacia tem 4 sessoes 0,1,2 e 3
        i = 0
        while 1:
            try:  # Checa se eh valido
                coluna0 = sheet.cell(i, 0).value  # """ Column and row must be at least 1, it'll break while -Lucas"""
            except:
                break

            if coluna0 != '':
                # Muda de sub-bacia e reseta sessao
                try:
                    if 'Sub-bacia' in coluna0:
                        subBacia = coluna0.split(":")[1].strip("\n")
                        subBacia = remover_acentos(subBacia).strip()
                        sessao = 0
                        dictPosto_Distr[subBacia] = {}
                        coluna0 = sheet.cell(i, 0).value  # Atualiza coluna0
                        i = i + 3
                except:
                    pass

                try:  # Troca de sessao toda vez que tem a palavra codigo
                    if 'digo' in coluna0:
                        sessao = sessao + 1
                        coluna0 = sheet.cell(i, 0).value  # atualiza coluna 0
                        i = i + 1  # Pula uma linha se tiver a palavra codigo
                except:
                    pass
                # cada sessao comeca quando a palavra "codigo" aparece..
                if coluna0 != '':
                    if sessao == 0:  # Reservatorio contido com pesos
                        try:
                            cod = int(sheet.cell(i, 0).value)
                            try:
                                dictPosto_Distr[subBacia]['posto'][cod] = []
                            except:
                                dictPosto_Distr[subBacia]['posto'] = {}
                                dictPosto_Distr[subBacia]['posto'][cod] = []
                            for j in range(12):
                                pesoMes = sheet.cell(i, j + 2).value
                                if pesoMes == '':
                                    pesoMes = 1
                                dictPosto_Distr[subBacia]['posto'][cod].append(pesoMes)
                        except:
                            pass

                    if sessao == 1:  # evapotranspiracao... nao eh usado
                        pass
                    if sessao == 2:  # Postos plu... nao eh usado
                        pass
                    if sessao == 3:  # postos flu a jusante com tempo de viagem
                        if sheet.cell(i, 6).value != '':
                            try:
                                nomeMontante = sheet.cell(i, 1).value
                                nomeMontante = remover_acentos(nomeMontante)
                                nomeJusante = sheet.cell(i, 2).value
                                nomeJusante = remover_acentos(nomeJusante)
                                tempoViagem = sheet.cell(i, 6).value
                                try:
                                    dictPosto_Montante_Viagem[nomeJusante.strip()][nomeMontante.strip()] = float(
                                        tempoViagem)
                                except:
                                    dictPosto_Montante_Viagem[nomeJusante.strip()] = {}
                                    dictPosto_Montante_Viagem[nomeJusante.strip()][nomeMontante.strip()] = float(
                                        tempoViagem)
                            except:
                                pass
            i = i + 1
    # Troca os nomes pelos codigos - Montante # '''Jusante - Lucas'''
    dictPosto_Montante_Viagem_Ajustado = {}

    for Nome_UHE in dictPosto_Montante_Viagem.keys():

        cod_uhe = Relacao_NomeUsina_PostoUHE(Nome_UHE)

        dictPosto_Montante_Viagem_Ajustado[cod_uhe] = {}
        for nome_montante in dictPosto_Montante_Viagem[Nome_UHE].keys():
            cod_flu_montante = Relacao_NomeSubBacia_PostoFlu(nome_montante)
            deltaT = dictPosto_Montante_Viagem[Nome_UHE][nome_montante]
            dictPosto_Montante_Viagem_Ajustado[cod_uhe][cod_flu_montante] = deltaT

    dictPosto_Montante_Viagem = dictPosto_Montante_Viagem_Ajustado  # [jusante][montante][tempo]
    # Troca os nomes pelos codigos - Distribuicao
    dictPosto_Distr_Aux = {}
    for nome_Subbacia in dictPosto_Distr.keys():
        dictAux = dictPosto_Distr[nome_Subbacia]
        cod_subbacia = Relacao_NomeSubBacia_PostoFlu(nome_Subbacia)
        dictPosto_Distr_Aux[cod_subbacia] = dictAux

    dictPosto_Distr = dictPosto_Distr_Aux
    return dictPosto_Distr, dictPosto_Montante_Viagem


def Leitura_CodFlu_SMAP(endereco, SMAPS):
    dictNomeArq_CodFlu = {}
    dictRegiaoPostoFluvDataVazao = {}
    for regiao in SMAPS:
        dictRegiaoPostoFluvDataVazao[regiao] = {}
        arquivo = endereco + '\\' + regiao + '\\ARQ_ENTRADA\\CASO.txt'
        enderecoRegiao = endereco + '\\' + regiao + '\\ARQ_ENTRADA'
        with open(arquivo, "r") as f:
            casoTXT = f.readlines()

        for line in casoTXT[1:]:
            PostoFlu = line.split('\'')[0].strip()
            PostoFlu = remover_acentos(PostoFlu)
            nomeFlu = line.split('\'')[1].strip()  # postos no CASO.txt
            nomeFlu = remover_acentos(nomeFlu)

            enderecoPostoFlu = enderecoRegiao + '\\' + PostoFlu + '.txt'
            with open(enderecoPostoFlu, "r") as filePostoFlu:
                # Obtendo a codificacao do posto
                linhasFlu = filePostoFlu.readlines()
                codFlu = linhasFlu[0].split("|")[0]
                codFlu = int(codFlu)  # Código do posto
                dictRegiaoPostoFluvDataVazao[regiao][codFlu] = {}  # Dicionario vazao por posto por regiao
                for linha in linhasFlu:
                    data = linha.split("|")[4]
                    data = data.split()[0]
                    data = data.split('-')
                    data = date(int(data[0]), int(data[1]), int(data[2]))
                    # data = Strdatetime_to_ExcelDate(data)
                    # data = xldate_to_datetime(data)
                    vaz = linha.split("|")[-1]

                    dictRegiaoPostoFluvDataVazao[regiao][codFlu][data] = float(vaz)

            dictNomeArq_CodFlu[PostoFlu] = int(codFlu)  # dic codigo por posto
    # retorna um dicionario de codigo por posto e outro dicionario de vazao por posto por regiao
    return dictNomeArq_CodFlu, dictRegiaoPostoFluvDataVazao


def remover_acentos(text):
    """
    Strip accents from input String.

    :param text: The input string.
    :type text: String.

    :returns: The processed String.
    :rtype: String.
    """
    try:
        text = np.unicode(text, 'utf-8')
    except (TypeError, NameError):  # unicode is a default on python 3
        pass
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore')
    text = text.decode("utf-8")
    return str(text)


def Strdatetime_to_ExcelDate(data):
    if '/' in data:
        char = '/'
    else:
        char = '-'

    listDt = data.split(char)
    if len(listDt[0]) == 4:
        ano = listDt[0]
        mes = listDt[1]
        dia = listDt[2]
    else:
        ano = listDt[2]
        mes = listDt[1]
        dia = listDt[0]
    dataAtual = datetime(int(ano), int(mes), int(dia))

    temp = datetime(1899, 12, 30)
    delta = (dataAtual - temp).days

    return delta


def xldate_to_datetime(xldate):
    temp = date(1899, 12, 30)
    delta = timedelta(days=xldate)
    return temp + delta


def Relacao_NomeUsina_PostoUHE(nomeUHE):
    if nomeUHE == 'P. COLOMBIA':
        cod = 12
    elif nomeUHE == 'MARIMBONDO':
        cod = 17
    elif nomeUHE == 'FURNAS':
        cod = 6
    elif nomeUHE == 'Sao Simao':
        cod = 33
    elif nomeUHE == 'ITAIPU':
        cod = 266
    elif nomeUHE == 'G. B. Munhoz':
        cod = 74
    #    '''URUGUAI'''
    elif nomeUHE == 'Campos Novos':
        cod = 216
    elif nomeUHE == 'Garibaldi':
        cod = 89
    else:
        print('erro- nao foi encontrado codigo para a usina com nome: ' + nomeUHE)
        sys.exit()
    return cod


def Relacao_NomeSubBacia_PostoFlu(nomeSubbacia):
    # Precisa estar IDENTICO ao arquivo de Configuracao !!!!!!!!!!!!!!!!!!!!!!!

    if nomeSubbacia == 'Serra do Facao':
        cod = 60035000
    elif nomeSubbacia == 'Agua Vermelha':
        cod = 61998080
    elif nomeSubbacia == 'Florida Estrada':
        cod = 64999999
    elif nomeSubbacia == 'Capivara':
        cod = 64516080
    elif nomeSubbacia == 'Itaipu':
        cod = 64918980
    elif (nomeSubbacia == 'Balsa Santa Maria') or (nomeSubbacia == 'Balsa Sta Maria'):
        cod = 64831000
    elif nomeSubbacia == 'Itumbiara':
        cod = 60610080
    elif nomeSubbacia == 'Funil':
        cod = 61146080
    elif nomeSubbacia == 'Uniao da Vitoria':
        cod = 65310001
    elif nomeSubbacia == 'Porto Taquara':
        cod = 64693000
    elif nomeSubbacia == 'Corumba IV':
        cod = 60444000
    elif nomeSubbacia == 'Jurumirim':
        cod = 64215080
    elif nomeSubbacia == 'Porto Colombia':
        cod = 61796080
    elif nomeSubbacia == 'Chavantes':
        cod = 64270080
    elif nomeSubbacia == 'Porto dos Buenos':
        cod = 61537000
    elif nomeSubbacia == 'Santa Clara':
        cod = 65824950
    elif nomeSubbacia == 'Canoas I':
        cod = 64345080
    elif nomeSubbacia == 'Capao Escuro':
        cod = 61787000
    elif nomeSubbacia == 'Ivinhema':
        cod = 64617000
    elif (nomeSubbacia == 'Jordao-Segredo') or (nomeSubbacia == 'Jordao - Segredo'):
        cod = 65826606
    elif nomeSubbacia == 'Furnas':
        cod = 61661000
    elif nomeSubbacia == 'Foz do Areia':
        cod = 65774403
    elif nomeSubbacia == 'Paraguacu':
        cod = 61425000
    elif nomeSubbacia == 'Passagem':
        cod = 61915000
    elif nomeSubbacia == 'Salto Caxias':
        cod = 65973500
    elif nomeSubbacia == 'Nova Ponte':
        cod = 60330080
    elif nomeSubbacia == 'Emborcacao':
        cod = 60160080
    elif nomeSubbacia == 'Camargos':
        cod = 61065080
    elif nomeSubbacia == 'Marimbondo':
        cod = 61941080
    elif nomeSubbacia == 'Rosana':
        cod = 64571080
    elif nomeSubbacia == 'Corumba I':
        cod = 60460000
    elif nomeSubbacia == 'Euclides da Cunha':
        cod = 61818080
    elif nomeSubbacia == 'Maua':
        cod = 64490080
    elif nomeSubbacia == 'Abaixo do Rio Verde':
        cod = 60805100
    elif nomeSubbacia == 'Sao Simao':
        cod = 60877080
    elif nomeSubbacia == 'Retiro Baixo':
        cod = 40865180
    elif nomeSubbacia == 'Tres Marias':
        cod = 40990080
    elif nomeSubbacia == 'Serra da Mesa':
        cod = 20920080
    #    '''URUGUAI'''
    elif nomeSubbacia == 'Barra Grande':
        cod = 70840080
    elif nomeSubbacia == 'Serra da Mesa':
        cod = ''
    elif nomeSubbacia == 'Campos Novos':
        cod = 71960080
    elif nomeSubbacia == 'Machadinho':
        cod = 72690081
    elif nomeSubbacia == 'Ita':
        cod = 73200080
    elif nomeSubbacia == 'Monjolinho':
        cod = 73500080
    elif nomeSubbacia == 'Foz do Chapeco':
        cod = 74040080
    elif nomeSubbacia == 'Quebra Queixo':
        cod = 73600580
    elif nomeSubbacia == 'Passo Sao Joao':
        cod = 75320100
    elif nomeSubbacia == 'Sao Francisco':
        cod = 44200000
    elif nomeSubbacia == 'Edgard de Souza':
        cod = 62335001
    elif nomeSubbacia == 'Queimado':
        cod = 42459080
    elif nomeSubbacia == 'Nova Avanhandava':
        cod = 62829580
    elif nomeSubbacia == 'Barra Bonita':
        cod = 62729080
    elif nomeSubbacia == 'Ibitinga':
        cod = 62790080
    elif nomeSubbacia == 'Sao Romao':
        cod = 43200000
    elif nomeSubbacia == 'Sao Francisco':
        cod = 44200000

    else:
        print('erro- nao foi encontrado codigo para a subbacia: ' + nomeSubbacia)
        sys.exit()

    return cod


def Leitura_Relacao_INC(path_relacao_inc, listRegiao):
    dictJusanteMontanteDelta = {}
    dictPostoBacia = {}
    try:
        with open(path_relacao_inc, "r") as f:
            reader = csv.reader(f, delimiter=";")
            first_line = 1
            for row in reader:
                if first_line == 1:
                    first_line = 0
                else:
                    bacia = row[1]
                    jusante = int(row[2])
                    if bacia in listRegiao:
                        try:
                            montante = int(row[3])
                        except:
                            montante = row[3]
                        deltaT = float(row[4])
                        try:
                            dictJusanteMontanteDelta[jusante][montante] = deltaT
                        except:
                            dictJusanteMontanteDelta[jusante] = {}
                            dictJusanteMontanteDelta[jusante][montante] = deltaT
                        # Identificador de bacia
                        dictPostoBacia[montante] = bacia


    except:
        print("Por favor, coloque o arquivo de relacao de INCs em: " + path_relacao_inc)
    return dictJusanteMontanteDelta, dictPostoBacia


def Leitura_Saida(listSMAPS, endereco, dictNomeArq_Cod):
    # Lista arquivos, devolve dict com {regiaoSMAP [posto,Arquivo]} e dictPostos com {regiaoSMAP:postos}
    dictArquivos, dictPostos, dictPostoCoord = Lista_Arquivos_Seleciona_Previsao(listSMAPS, endereco)
    dictPostoDataVazao = {}
    dictRegiaoPostoDataVazao = {}

    for regiao in listSMAPS:
        dictPostoDataVazao = {}
        for arquivo in dictArquivos[regiao]:
            file = open(arquivo[1], 'r')
            posto = arquivo[0]
            dictDataVazao = {}
            for line in file:
                if 'Data' not in line:
                    data = line.split()[0]
                    data = Strdatetime_to_ExcelDate(data)
                    data = xldate_to_datetime(data)
                    vazao = line.split()[1]
                    dictDataVazao[data] = float(vazao)
                    cod = dictNomeArq_Cod[posto]
            dictPostoDataVazao[cod] = dictDataVazao
        dictRegiaoPostoDataVazao[regiao] = dictPostoDataVazao

    return dictRegiaoPostoDataVazao


def Lista_Arquivos_Seleciona_Previsao(listSMAPS, endereco):
    dictPostos = {}
    dictArquivoTxt = {}
    dictPostoCoord = {}
    for regiao in listSMAPS:
        dictArquivoTxt[regiao] = os.listdir(endereco + '\\' + regiao + '\\Arq_Saida')
        setPostos = set()
        setTiposPrevisoes = set()
        # Guarda os tipos de postos
        dictPostoCoord[regiao] = {}
        for arquivoTxt in dictArquivoTxt[regiao]:
            #if 'PRECMEDIA_ETA40.GEFS' in arquivoTxt:
            if 'PMEDIA' in arquivoTxt:
                # Ajusta nome dos postos
                nomePosto = arquivoTxt.partition("_")[0]
                setPostos.add(nomePosto)
                # Ajusta tipos de previsão
                auxTxt = arquivoTxt.partition("_")[2]
                auxTxt = auxTxt.partition(".txt")[0]
                setTiposPrevisoes.add(auxTxt)

                dictPostoCoord[regiao][nomePosto] = Le_Coordenadas(
                    endereco + '\\' + regiao + '\\ARQ_ENTRADA\\' + nomePosto + '_PRECMEDIA_ETA40.GEFS.txt')


        dictPostos[regiao] = list(setPostos)

    # Escolhe previsao
    listTiposPrevisoes = list(setTiposPrevisoes)
    PREVISAO = listTiposPrevisoes[0]
    dictArquivosPrevisao = {}
    for regiao in listSMAPS:
        enderecoRegiao = endereco + '\\' + regiao + '\\Arq_Saida'
        listArquivosPrevisao = []
        for arquivoTxt in dictArquivoTxt[regiao]:
            if PREVISAO in arquivoTxt:
                enderecoRegiaoArquivo = enderecoRegiao + '\\' + arquivoTxt
                for posto in dictPostos[regiao]:
                    if posto in arquivoTxt:
                        postoAtual = posto
                listArquivosPrevisao.append([postoAtual, enderecoRegiaoArquivo])
        dictArquivosPrevisao[regiao] = listArquivosPrevisao
    return dictArquivosPrevisao, dictPostos, dictPostoCoord


def Le_Coordenadas(arq):
    listLongLat = []
    with open(arq) as txt:
        list_arq = txt.readlines()
        num = list_arq[0]
        for i in range(int(num)):
            long_lat = list_arq[i + 1].strip('\n')
            long_lat = long_lat.split()
            listLongLat.append(long_lat)
    return listLongLat


def Compila_Entrada_Saida(dictRegiaoPostoFluvDataVazao_ENT, dictRegiaoPostoFluvDataVazao_SAI):
    dictRegiaoPostoFluvDataVazao = {}
    dictRegiaoPostoFluvDataVazao = dictRegiaoPostoFluvDataVazao_ENT
    for regiao in dictRegiaoPostoFluvDataVazao_SAI.keys():
        for posto in dictRegiaoPostoFluvDataVazao_SAI[regiao].keys():
            for data in dictRegiaoPostoFluvDataVazao_SAI[regiao][posto].keys():
                vaz = dictRegiaoPostoFluvDataVazao_SAI[regiao][posto][data]
                dictRegiaoPostoFluvDataVazao[regiao][posto][data] = vaz
    return dictRegiaoPostoFluvDataVazao


def Calcula_VNI(dictPostoFluvDataVazao, dictPosto_Distr, dictPosto_Montante_Viagem):
    #    print dictPosto_Distr.keys()
    #    print dictPostoFluvDataVazao.keys()                                                #
    #    sys.exit()                                                                         #
    dictUHEDataVazao = {}  #
    for posto in dictPostoFluvDataVazao.keys():  #
        listData = sorted(dictPostoFluvDataVazao[posto].keys())[10:]  #
        try:
            listUHE = dictPosto_Distr[posto]['posto'].keys()
        except:
            listUHE = []
        for uhe in listUHE:

            flag_inc = 0
            dictUHEDataVazao[uhe] = {}
            listPesos = dictPosto_Distr[posto]['posto'][uhe]
            for data in listData:
                Indice_mes = data.month - 1
                peso = listPesos[Indice_mes]
                vazao = dictPostoFluvDataVazao[posto][data]
                vazaoTot = vazao * peso
                # Adiciona incremental se houver
                try:
                    dictPosto_Montante_Viagem[uhe]
                    flag_inc = 1
                except:
                    pass
                if flag_inc == 1:
                    vazaoINC = Calcula_INC(uhe, dictPostoFluvDataVazao, dictPosto_Montante_Viagem, data)
                    vazaoTot += vazaoINC

                dictUHEDataVazao[uhe][data] = vazaoTot

    return dictUHEDataVazao


def Calcula_INC(posto_uhe, dictPostoFluvDataVazao, dictPosto_Montante_Viagem, data):
    somaINC = 0
    for posto_montante in dictPosto_Montante_Viagem[posto_uhe].keys():
        deltaT = dictPosto_Montante_Viagem[posto_uhe][posto_montante]
        if deltaT < 24:
            vazao0d = dictPostoFluvDataVazao[posto_montante][data]
            data1d = data - timedelta(days=1)
            vazao1d = dictPostoFluvDataVazao[posto_montante][data1d]

            somaINC += ((24 - deltaT) * vazao0d + deltaT * vazao1d) / 24
        else:
            data1d = data - timedelta(days=1)
            data2d = data - timedelta(days=2)
            vazao1d = dictPostoFluvDataVazao[posto_montante][data1d]

            vazao2d = dictPostoFluvDataVazao[posto_montante][data2d]

            somaINC += ((48 - deltaT) * vazao1d + (deltaT - 24) * vazao2d) / 24

    return somaINC


def Calcula_VNA(dictUHEDataVazao, dictRelacao_INC):
    dictUHEDataVazaoTot = {}
    for posto in dictRelacao_INC.keys():
        dictUHEDataVazaoTot[posto] = {}
        listDatas = sorted(dictUHEDataVazao[posto].keys())[5:]
        for data in listDatas:
            vazaoT = 0
            vazaoT = Vazao_Total(posto, data, vazaoT, dictUHEDataVazao, dictRelacao_INC)
            dictUHEDataVazaoTot[posto][data] = vazaoT

    return dictUHEDataVazaoTot


def Vazao_Total(posto, data, vazaoT, dictUHEDataVazao, dictRelacao_INC):
    vazaoINC = 0
    #    try
    #    print dictRelacao_INC
    #    sys.exit()
    listMontante = dictRelacao_INC[posto].keys()
    
    for posto_montante in listMontante:
        if posto_montante != '':
            vazaoINC += Vazao_Total(posto_montante, data, vazaoINC, dictUHEDataVazao, dictRelacao_INC)

        ########### delta temporal de vazoes nao implementado!!! ######################
        #            deltaT = dictRelacao_INC[posto][posto_montante]
        #            if deltaT < 24:
        #                vazao0d = dictUHEDataVazao[posto_montante][data]
        #                data1d = data -timedelta(days=1)
        #                vazao1d = dictUHEDataVazao[posto_montante][data1d]
        #
        #                vazaoINC+= ((24-deltaT)*vazao0d +deltaT*vazao1d)/24
        #            else:
        #                data1d = data -timedelta(days=1)
        #                data2d = data -timedelta(days=2)
        #                vazao1d = dictUHEDataVazao[posto_montante][data1d]
        #                vazao2d = dictUHEDataVazao[posto_montante][data2d]
        #
        #                vazaoINC+=((48 -deltaT)*vazao1d +(deltaT-24)*vazao2d)/24
        ###############################################################################
        else:
            pass
    vazaoT += vazaoINC
    return vazaoT


def Escreve_Planilha_Resultados(listRegiao, dictUHEDataVazao, enderecoResultados, dataRodada, listDatas,
                                dictPostoBacia):
    time.sleep(4)
    correlacao = []
    CodigoRV = Detecta_RV_Atual(dataRodada)
    if 'Grande' in listRegiao:
        # Código - Cod Usina - resultados.xlsxVertical
        correlacao = correlacao + [['GRCAMA', 1, 4],
                                   ['GRITUT', 2, 88],
                                   ['GRFUNI', 211, 928],
                                   ['GRFURN', 6, 172],
                                   ['GRMMOR', 7, 256],
                                   ['GRLCBA', 8, 340],
                                   ['GRJAGU', 9, 424],
                                   ['GRIGAR', 10, 508],
                                   ['GRVGRA', 11, 592],
                                   ['GRPCOL', 12, 676],
                                   ['GRMARI', 17, 760],
                                   ['GRAGVL', 18, 844],
                                   ['PDCACO', 14, 1012],
                                   ['PDECUN', 15, 1180],
                                   ['PDLIMO', 16, 1096]]

    if 'Paranapanema' in listRegiao:
        correlacao = correlacao + [['PPJURU', 47, 2356],  # Jurumirim
                                   ['PPPJU', 48, 2440],  # Piraju
                                   ['PPCHAV', 49, 2524],
                                   ['PPSGCS', 50, 2608],  # Salto Grande
                                   ['PPCNO2', 51, 2692],  # Canoas 2
                                   ['PPCNO1', 52, 2776],
                                   ['PPUHMU', 57, 2860],  # MAUA
                                   ['PPCAPI', 61, 2944],  # CAPIVARA
                                   ['PPTAQU', 62, 3028],  # TAQUARUÇU
                                   ['PPROSA', 63, 3112],
                                   ['PPORI', 249, 3196]]  # Ourihos
    if 'Paranaiba' in listRegiao:
        correlacao = correlacao + [['PNUSBL', 22, 1264],  # BATALHA
                                   ['PNUCB3', 23, 1348],  # CORUMBA3
                                   ['PREMBO', 24, 1432],  # EMBORCACAO
                                   ['PRNPON', 25, 1516],  # NOVAPONTE
                                   ['PNCPB2', 28, 1600],  # C. BRANCO2
                                   ['PRITUM', 31, 1684],  # ITMBIARA
                                   ['PRCDOU', 32, 1768],  # C. DOURADA
                                   ['PRITUM', 33, 1852],  #
                                   ['PNUCB4', 205, 1936],  # CORUMBA 4
                                   ['ARMIRA', 206, 2020],  # MIRANDA
                                   ['PNCPB1', 207, 2104],  # C.BRANCO 1
                                   ['COCORU', 209, 2188],  # CORUMBA
                                   ['PNUHSF', 251, 2272]]  # S. DO FALCAO
    if 'Iguacu' in listRegiao:
        correlacao = correlacao + [['IGCSCL', 71, 3280],
                                   ['IGCFUN', 72, 3364],
                                   ['IGCJOR', 73, 3448],
                                   ['IGCGBM', 74, 3532],
                                   ['IGCSEG', 76, 3616],
                                   ['IGCSST', 77, 3700],
                                   ['IGCSOS', 78, 3784],
                                   ['IGCSCX', 222, 3868]]
    if 'Itaipu' in listRegiao:
        correlacao = correlacao + [['ITAIPU INC', 266, 3952]]
    if 'SaoFrancisco' in listRegiao:
        correlacao = correlacao + [['SFRBAI', 155, 4036],
                                   ['SFTMAR', 156, 4120],
                                   ['SFTQUEI', 158, 5968]]

    if 'Tocantins' in listRegiao:
        correlacao = correlacao + [['TOSERM', 270, 4204]]

    if 'Uruguai' in listRegiao:
        correlacao = correlacao + [['URGARI', 89, 4288],
                                   ['URCMNV', 216, 4372],
                                   ['URBRGR', 215, 4456],
                                   ['URMACH', 217, 4540],
                                   ['URITA', 92, 4624],
                                   ['URPSFD', 93, 4708],
                                   ['URMONJ', 220, 4792],
                                   ['URFCHP', 94, 4876],
                                   ['URQBQX', 286, 4960],
                                   ['URSJOS', 102, 5044],
                                   ['URPSJO', 103, 5128]]
    if 'Tiete' in listRegiao:
        correlacao = correlacao + [['TTBPED', 119, 5212],  # Billings + Pedreira
                                   ['TTEDSO', 161, 5296],  # Edgar Souza+ Pinheiros
                                   ['TTGUAP', 117, 5380],  # Guarapiranga
                                   ['TTPONV', 160, 5464],  # Ponte Nova
                                   ['TTBBON', 237, 5548],  # Barra Bonita
                                   ['TTBARR', 238, 5632],  # Bariri
                                   ['TTIBIT', 239, 5716],  # Ibitinga
                                   ['TTNAVA', 242, 5800],  # Nova Avanhandava
                                   ['TTPROM', 240, 5884]]  # Promissao

    print('Escrevendo resultados em: ' + enderecoResultados)

    #    w2 = openpyxl.load_workbook(enderecoResultados,data_only=True)
    #    wsRes = w2.worksheets[3]
    wb = load_workbook(enderecoResultados)
    sheet = wb.get_sheet_by_name('TOT')

    posResDataRodada_H = 7
    posResDataPrev_H = 8
    posResVazao_H = 9
    posVerificado_H = 10
    posResCodigoRV_H = 11

    #    print( "Regiao: " +regiao)
    dictPostoDataVazao = {}
    for posto_apoio_resul in correlacao:
        nomePosto = posto_apoio_resul[0]
        #        bacia = dictPostoBacia[nomePosto]
        if 1:
            dictPostoDataVazao[nomePosto] = {}
            print(nomePosto)
            # time.sleep(1)##########################################   Retornar
            codPosto = posto_apoio_resul[1]
            posRes_V = posto_apoio_resul[2]
            for dataPrev in listDatas:  # Usa-se o valor 84 para 12 semanas, alterar aqui caso queira alterar

                # Campo verificado
                # time.sleep(0.02)#######################################    Retornar
                #            print(dataPrev)

                if dataPrev > dataRodada:
                    verificado = 't'
                else:
                    verificado = 'f'
                # Escreve na planilha resultados

                vazao = dictUHEDataVazao[codPosto][dataPrev]
                sheet.cell(row=posRes_V, column=posResDataRodada_H).value = dataRodada
                sheet.cell(row=posRes_V, column=posResDataPrev_H).value = dataPrev
                sheet.cell(row=posRes_V, column=posResVazao_H).value = vazao
                sheet.cell(row=posRes_V, column=posVerificado_H).value = verificado
                sheet.cell(row=posRes_V, column=posResCodigoRV_H).value = CodigoRV
                time.sleep(0.02)
                posRes_V = posRes_V + 1

    #    dataHoje = date.today().
    strData = date.today().strftime("%Y_%m_%d")
    enderecoResultadosNovo = 'Arq_saida\\' + strData + 'resultados.xlsx'
    wb.save(filename=enderecoResultadosNovo)

    print('-Planilha resultados.xlsx Atualizada com SUCESSO!')

    return dictPostoDataVazao


def Detecta_RV_Atual(dataRodada):
    dataHoje = dataRodada
    data1doMes = dataHoje.replace(day=1)
    dataAux = data1doMes
    rv = 0
    while dataAux < dataHoje:
        dataAux = dataAux + timedelta(days=1)
        if dataAux.weekday() == 5:
            rv = rv + 1

    while dataAux.weekday() != 4:
        dataAux = dataAux + timedelta(days=1)

    if dataHoje.month != dataAux.month:
        rv = 0
        mes = int(dataHoje.month) + 1
        ano = int(dataAux.year)  # Caso vire o ano...
    else:
        mes = int(dataHoje.month)
        ano = int(dataAux.year)

    strRVMes = 'RV' + str(rv) + str(ano) + str(mes).zfill(2)

    return strRVMes


def Cria_Pasta(path):
    if not os.path.isdir(path):
        os.makedirs(path)


def Edita_Entrada(listSMAPS, endereco, dataRodada,
                  diasPrevisao):  # #Data rodada e data previsão estão invertidas... (listSMAOS, endereco,
    # DiasPrevisao, Data hj)
    print('Editando entrada')
    dictArquivos = {}
    for regiao in listSMAPS:

        enderecoEntrada = endereco + '\\' + regiao + '\\Arq_Entrada'
        ### EDITA MODELO_PRECIPITACAO ######
        with open(enderecoEntrada + '\\MODELOS_PRECIPITACAO.txt', 'w') as modelo:
            textoModelos = "1\nPMEDIA_ORIG\n"
            #textoModelos = "2\nPMEDIA_ORIG\nPMEDIA"#PRECZERADA_\n"
            #textoModelos = "1                                                 'Quantidade de modelos de previsão de " \
                          # "precipitação\nPRECMEDIA_ETA40.GEFS "

            modelo.write(textoModelos)

        ####################################
        dictArquivos[regiao] = os.listdir(enderecoEntrada)
        # Guarda os tipos de postos
        # Edita os txts para a nova rodada e com os novos dias
        # Leitura da dataInicial do modelo
        for arquivoTxt in dictArquivos[regiao]:
            if 'INICIALIZACAO' in arquivoTxt:
                dataInit = Edita_Entrada_Caso(enderecoEntrada + '\\' + arquivoTxt, dataRodada, diasPrevisao)
                # Ajusta nome dos postos
    return dataInit


def Edita_Entrada_Caso(arquivo, diasPrevisao, dataRodada):
    #        strDataRodada = str(dataRodada.day).zfill(2)+'/'+str(dataRodada.month).zfill(2)+'/'+str(dataRodada.year)
    with open(arquivo, "r") as f:
        old = f.read()

        #        old = stDataRodada+old[10:] ### Para alterar a data, mudar a variavel strDataRodada
        new = old[:120] + str(diasPrevisao) + old[122:]  # Posicoes devem ser fixas!!!!!!
    time.sleep(0.02)

    with open(arquivo, 'w') as file:
        file.write(new)

    strDt = new[:10]
    listStrDt = strDt.split('/')
    dia = listStrDt[0]
    mes = listStrDt[1]
    ano = listStrDt[2]
    dataInit = date(int(ano), int(mes), int(dia))
    time.sleep(0.3)
    return dataInit


def main(path, data, dias_previsao):
    path_SMAP = path
    SMAPS = ['Tiete', 'SaoFrancisco', 'Uruguai', 'Itaipu', 'Grande', 'Paranaiba', 'Paranapanema', 'Iguacu', 'Tocantins']
    path_conf_SMAP = 'base\\dev\\Configuracao_SMAP.xlsx'
    path_relacao_inc = 'base\\dev\\Relacao_INC.csv'
    enderecoResultados = 'base\\dev\\resultados.xlsx'
    diasPrevisao = dias_previsao
    dataHoje = data
    dataInit = Edita_Entrada(SMAPS, path_SMAP, diasPrevisao, dataHoje)

    print("##\n3c) Aplicando configuracoes de postos e vazoes!")

    dictNomeArq_Cod, dictRegiaoPostoFluvDataVazao_ENT = Leitura_CodFlu_SMAP(path_SMAP, SMAPS)
    dictPosto_Distr, dictPosto_Montante_Viagem = Leitura_Configuracao(path_conf_SMAP)
    dictRelacao_INC, dictPostoBacia = Leitura_Relacao_INC(path_relacao_inc, SMAPS)

    dictRegiaoPostoFluvDataVazao_SAI = Leitura_Saida(SMAPS, path_SMAP, dictNomeArq_Cod)
    dictRegiaoPostoFluvDataVazao = Compila_Entrada_Saida(dictRegiaoPostoFluvDataVazao_ENT,
                                                         dictRegiaoPostoFluvDataVazao_SAI)

    # Retirando regioes do dict
    dictPostoFluvDataVazao = {}
    for regiao in dictRegiaoPostoFluvDataVazao.keys():

        for posto in dictRegiaoPostoFluvDataVazao[regiao].keys():
            
            dictPostoFluvDataVazao[posto] = dictRegiaoPostoFluvDataVazao[regiao][posto]

    dictUHEDataVazao = Calcula_VNI(dictPostoFluvDataVazao, dictPosto_Distr, dictPosto_Montante_Viagem)

    dictUHEDataVazTot = Calcula_VNA(dictUHEDataVazao, dictRelacao_INC)

    # listDatas = []
    # dt = dataInit
    # for i in range(diasPrevisao-1): # ESSE -1 AQUI...
    # dt = dt +timedelta(days=1)
    # listDatas.append(dt)
    # Coloquei a partir daqui Erinaldo
    listDatas = []
    dt = dataInit
    print('here')
    while dt.weekday() != 5:  # ESSE -1 AQUI...
        dt = dt + timedelta(days=-1)
        listDatas.append(dt)
    # Dias Previsao
    dt = dataInit
    listDatas.append(dt)
    for i in range(diasPrevisao - 1):
        dt += timedelta(days=1)
        listDatas.append(dt)
    listDatas = sorted(listDatas)

    Escreve_Planilha_Resultados(SMAPS, dictUHEDataVazTot, enderecoResultados, dataInit, listDatas,
                                dictPostoBacia)



if __name__ == '__main__':
    main()